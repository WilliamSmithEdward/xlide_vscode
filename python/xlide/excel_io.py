"""Excel cell read/write via openpyxl."""
from __future__ import annotations

import contextlib
import io
import json as _json
from typing import Any

import openpyxl
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

from xlide.vba_io import get_modules_and_protection_info as _get_modules_and_protection_info


def _parse_cell(ref: str) -> tuple[int, int]:
    """Return (row, col) 1-based integers from a cell reference like 'B3'."""
    col_str, row = coordinate_from_string(ref)
    return row, column_index_from_string(col_str)


def _sheet_summaries(wb: Any) -> list[dict[str, Any]]:
    """Return [{name, dimensions}] for every worksheet in an open workbook."""
    return [
        {"name": ws.title, "dimensions": ws.dimensions or ""}
        for ws in wb.worksheets
    ]


def get_workbook_info(*, path: str) -> dict[str, Any]:
    """Return a combined summary: VBA modules, sheet names/dimensions, named ranges, and protection state."""
    vba_info = _get_modules_and_protection_info(path=path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_vba=True)
    try:
        sheets = _sheet_summaries(wb)
        named_ranges = [
            {"name": nr.name, "ref": nr.attr_text}
            for nr in wb.defined_names.definedName
        ]
    finally:
        wb.close()
    return {
        "sheets": sheets,
        "namedRanges": named_ranges,
        **vba_info,
    }


def list_sheets(*, path: str) -> dict[str, Any]:
    """Return the sheet names and used dimensions for every worksheet."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_vba=True)
    try:
        sheets = _sheet_summaries(wb)
    finally:
        wb.close()
    return {"sheets": sheets}


def _read_range(path: str, sheet: str, range: str, *, data_only: bool) -> dict[str, Any]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=data_only, keep_vba=True)
    try:
        ws = wb[sheet]
        data: list[list[Any]] = [
            [cell.value for cell in row] for row in ws[range]
        ]
    finally:
        wb.close()
    return {"data": data}


def read_cells(*, path: str, sheet: str, range: str) -> dict[str, Any]:
    """Return {data: [[...]]} for a cell range in A1 notation."""
    return _read_range(path, sheet, range, data_only=True)


def read_formulas(*, path: str, sheet: str, range: str) -> dict[str, Any]:
    """Return {data: [[...]]} with raw formula strings (not computed values)."""
    return _read_range(path, sheet, range, data_only=False)


def write_cells(
    *, path: str, sheet: str, startCell: str, data: list[list[Any]]
) -> dict[str, Any]:
    """Write a 2-D array of values starting at startCell and save."""
    start_row, start_col = _parse_cell(startCell)

    wb = openpyxl.load_workbook(path, keep_vba=True)
    try:
        ws = wb[sheet]
        for r_offset, row in enumerate(data):
            for c_offset, value in enumerate(row):
                ws.cell(
                    row=start_row + r_offset,
                    column=start_col + c_offset,
                    value=value,
                )
        wb.save(path)
    finally:
        wb.close()
    return {"ok": True}


def run_openpyxl(
    *, path: str, code: str, save: bool = True
) -> dict[str, Any]:
    """Execute arbitrary openpyxl code with the workbook open.

    The code runs in a namespace containing:
      - ``wb``       -- the open openpyxl Workbook
      - ``openpyxl`` -- the openpyxl module
      - ``json``     -- the json stdlib module
      - ``result``   -- initialised to None; assign the return value here

    If *save* is True (default) and the code does not raise, the workbook
    is saved back to *path*.

    Returns ``{"result": <value of result>, "stdout": "<captured output>"}}``.
    """
    wb = openpyxl.load_workbook(path, keep_vba=True)
    namespace: dict[str, Any] = {
        "wb": wb,
        "openpyxl": openpyxl,
        "json": _json,
        "result": None,
    }
    buf = io.StringIO()
    try:
        with contextlib.redirect_stdout(buf):
            exec(compile(code, "<xlide_openpyxl>", "exec"), namespace)  # noqa: S102
        if save:
            wb.save(path)
    finally:
        wb.close()
    return {"result": namespace.get("result"), "stdout": buf.getvalue()}
